import openpyxl as op

wb = op.load_workbook("test.xlsx")

ws = wb['무']

ws.cell(1,2).value = "입력테스트1"
ws['C1'].value = "입력테스트2"

wb.save("test.xlsx")